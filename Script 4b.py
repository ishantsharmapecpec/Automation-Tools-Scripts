# -*- coding: utf-8 -*-
"""
Created on Thu Aug  1 11:03:09 2024

@author: ISTM
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import copy

import os

EXCEL_FILE = r"\\rammdhfile02\Filery Supplementary Folders\Buildings\2023\RUK2023N00283 - HS2 Curzon Station CAT III Checking\4 Delivery\Geo\11 Analyses\WP05\Stiffness matrix\Stiffness MatrixV33.xlsm"

# Load Excel file
df = pd.read_excel(EXCEL_FILE, sheet_name='Repute Input', engine='openpyxl')

def copy_range(src_ws, dest_ws, min_row, min_col, max_row, max_col, start_row, start_col):
    # Copy the cell values and styles
    for i, row in enumerate(src_ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col), start=start_row):
        for j, cell in enumerate(row, start=start_col):
            dest_cell = dest_ws.cell(row=i, column=j)
            
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                dest_cell.value = cell.value

            if cell.has_style:
                dest_cell.font = copy.copy(cell.font)
                dest_cell.border = copy.copy(cell.border)
                dest_cell.fill = copy.copy(cell.fill)
                dest_cell.number_format = copy.copy(cell.number_format)
                dest_cell.protection = copy.copy(cell.protection)
                dest_cell.alignment = copy.copy(cell.alignment)

    # Copy merged cells
    '''for merged_cell_range in src_ws.merged_cells.ranges:
        min_r, min_c, max_r, max_c = merged_cell_range.bounds
        print(min_r, min_c, max_r, max_c)
        if min_row <= min_r <= max_row and min_col <= min_c <= max_col:
            dest_ws.merge_cells(
                start_row=min_r - min_row + start_row,
                start_column=min_c - min_col + start_col,
                end_row=max_r - min_row + start_row,
                end_column=max_c - min_col + start_col,'''
            

min_row, min_col = 4, 2
max_row, max_col = 300, 16
start_row, start_col = 4, 2

dest_dir=r"\\rammdhfile02\Filery Supplementary Folders\Buildings\2023\RUK2023N00283 - HS2 Curzon Station CAT III Checking\4 Delivery\Geo\11 Analyses\WP05\Repute results for WP5\Pile group report\SET SLS CHAR\SLS CHAR 06Dec2024 for WP07"
src_dir=r"\\rammdhfile02\Filery Supplementary Folders\Buildings\2023\RUK2023N00283 - HS2 Curzon Station CAT III Checking\4 Delivery\Geo\11 Analyses\WP05\Repute results for WP5\Extracted spread sheet from Repute\SET SLS CH\SLS CH WP07 DATA EXTRCAED"


# Load workbooks and worksheets


for file_index in range(9, 363): #236
    name=str(df.iloc[file_index, 20])
    xlsx_file = name[:-3]+'xlsx'
    print(f"Processing file: {xlsx_file}")

    # Search for the XML file in the directory
    for root_dir, dirs, files in os.walk(dest_dir):
        if xlsx_file in files:
            print(f"Found file: {xlsx_file}")
            break
    else:
        print("Empty cell in excel spread sheet")
        continue
    


    for i in range(1,7):
        
        src_file=f"\{name[:-4]}"+f"_LC{i}.xlsx"
        src_wb = openpyxl.load_workbook(src_dir+src_file)
        
        try: 
            src_name=f'{name[:-4]}'+f'_LC{i}'
            src_ws = src_wb[src_name]
            
        except KeyError:
            src_name=f'{name[:-4]}'+f'_L'
            src_ws = src_wb[src_name]
                
        dest_file=f"\{xlsx_file}"
        dest_wb = openpyxl.load_workbook(dest_dir+dest_file)
        dest_ws = dest_wb[f'LC{i}']  # Replace 'Sheet1' with your worksheet name
        copy_range(src_ws, dest_ws, min_row, min_col, max_row, max_col, start_row, start_col)
        dest_wb.save(dest_dir+dest_file)
        print(f'LC{i} out of 6 sheet has been populated')
    
    # Define the range to copy
    
    
    # Copy the range
    #copy_range(src_ws, dest_ws, min_row, min_col, max_row, max_col, start_row, start_col)
    
    
    
    
    
    # Save the destination workbook
# dest_wb.save('destination.xlsx')